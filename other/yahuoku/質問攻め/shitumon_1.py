'''
------------------------------
使い方


------------------------------
更新履歴

前
	ihan_sinkoku_1 だと、読み込みエラーが発生して止まるので
	ihan_sinkoku_2
	はtryとwhileで正常終了するまでループするようにした

2019年8月28日
	auクーポンの場合は出品数が多いので
	ihan_sinkoku_3
	でページにまたがってオークションID取得できるようにした
	------------------------------
	エラーが表示されたときに、forで中断したところから再開できるようにした



------------------------------
'''


import poplib
from datetime import datetime
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from my_module import my_function, mail_receiving_pop3
import random
import chromedriver_binary
import time
import openpyxl
import win32com
import xl
import re
import subprocess
import chardet
from my_module import trashmail_mass_production
from selenium.common.exceptions import NoSuchElementException



Excel = 'DATA.xlsx'
wbpx = openpyxl.load_workbook(Excel)
wbxl = xl.Workbook(Excel)
# シート指定
sheet = wbpx['IDPASS']

# 違反申告するオークションID
auk_id = []

auk_comment = '今から即対応可能ですか？'


driver = webdriver.Chrome()
driver.set_window_size(1015, 515)

# ヤフオク商品ページに移動
# au クーポン (ID auID auid) のページ
page_url = "https://auctions.yahoo.co.jp/search/search?p=au+%E3%82%AF%E3%83%BC%E3%83%9D%E3%83%B3+%28ID+auID+auid%29&auccat=2084048817%2C23960%2C2084044317&va=au+%E3%82%AF%E3%83%BC%E3%83%9D%E3%83%B3&vo=ID+auID+auid&exflg=1&b=1&n=100&mode=2&f=0x4&rewrite_category=0"
driver.get(page_url)
print(my_function.browser_load_wait(driver))
# オークションID取得
soup = BeautifulSoup(driver.page_source, "html.parser")
for tag in soup.find_all('a', class_= "Product__imageLink"):
	# auction/ から始まって 最後までを切り取り
	tag = re.search('auction\/(.*$)', str(tag.get("href"))).group(1)
	auk_id.append(tag)
print(len(auk_id), "個のオークションを取得")

# ここからfor?
# for
# ログイン
url = "https://login.yahoo.co.jp/config/login"
driver.get(url)
print(my_function.browser_load_wait(driver))
# ExcelからIDパスワード取得
yahoo_id = sheet.cell(row=3, column=2).value
yahoo_pass = sheet.cell(row=3, column=3).value
print( 'yahoo_id=', yahoo_id, 'yahoo_pass=', yahoo_pass, 'でログイン')
driver.find_element_by_xpath('//*[@id="username"]').send_keys(yahoo_id)
driver.find_element_by_xpath('//*[@id="btnNext"]').click()
time.sleep(1)
driver.find_element_by_xpath('//*[@id="passwd"]').send_keys(yahoo_pass)
driver.find_element_by_xpath('//*[@id="btnSubmit"]').click()
time.sleep(1)
# print(my_function.browser_load_wait(driver))

# 質問する
while True:
	try:
		for count, auk_id_fa in enumerate(auk_id):
			while True:
				url = "https://auctions.yahoo.co.jp/jp/show/qanda?aID=" + auk_id_fa
				driver.get(url)
				my_function.browser_load_wait(driver)
				report_soup = BeautifulSoup(driver.page_source, "html.parser")
				if re.search('未回答の質問はありません。', str(report_soup)):
					print(count+1, "/", len(auk_id), "自分のオークションなのでスキップ")
					break
				elif re.search('質問はできません', str(report_soup)):
					print("ブロックされた")
				driver.find_element_by_xpath('//*[@id="comment"]').send_keys(auk_comment + ' ' + str(time.time()))
				driver.find_element_by_xpath('//*[@id="modFormSbt"]/div[1]/input[1]').click()
				my_function.browser_load_wait(driver)
				driver.find_element_by_xpath('//*[@id="modFormSbt"]/div[1]/table/tbody/tr/td[1]/form/input[1]').click()
				my_function.browser_load_wait(driver)
				print(auk_id[count], '質問', count+1, "回目")
	except NoSuchElementException as err:
		print('エラー', err)
	else:
		break

# 終了処理
print("正常終了")
driver.quit()




