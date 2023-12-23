import poplib
from datetime import datetime
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from my_module import my_function, mail_receiving_pop3
import random
import chromedriver_binary
import time
import openpyxl
import win32com
import xl
import re
import subprocess
import chardet
from my_module import trashmail_mass_production
from selenium.webdriver.common.alert import Alert



# ------------------------------

'''
Excel = 'aucoupon_data.xlsx'
wbpx = openpyxl.load_workbook(Excel)
wbxl = xl.Workbook(Excel)
sheet = wbpx['input_data']
mrow = sheet.max_row
mcol = sheet.max_column
for fa in range(mrow-1):
	# メルアド取得
	print(wbxl.get("B" + str(2 + fa)).get())
	# PASSをランダムに作成して記入
	# 8桁以上英数記号
	wbxl.get("C" + str(2+fa)).set(my_function.pass_gen(20))
'''
# ------------------------------

# ------------------------------

'''
pop3_serv = 'pop.mail.yahoo.co.jp'
pop3_user = 'trashmail_receive_001'
pop3_pass = 'j50fBU1PvH'

cli = poplib.POP3(pop3_serv)
cli.user(pop3_user)
cli.pass_(pop3_pass)
# 確認コードが記載されているメルを受信する前のメルの数を取得
mail_num_max = cli.stat()[0]
print('受信前のメル数', mail_num_max)

msg = mail_receiving_pop3.fetchmail(cli, mail_num_max)
print(msg[2])
'''



'''
捨てメアド を操作する
'''
driver = webdriver.Chrome()
driver.set_window_size(1015, 515)
url = "https://m.kuku.lu/index.php?pagemode_login=1&noindex=1"
driver.get(url)
print(my_function.browser_load_wait(driver))
# ------------------------------
# ログイン動作
user_number = '5917367876'
user_password = '629927'
# 他のアカウントにログイン を押す
driver.find_element_by_xpath('//*[@id="link_loginform"]').click()
time.sleep(0.5)
driver.find_element_by_xpath('//*[@id="user_number"]').send_keys(user_number)
driver.find_element_by_xpath('//*[@id="user_password"]').send_keys(user_password)
# ログイン を押す
driver.find_element_by_xpath("//input[@type='button'][@value='ログイン']").click()
time.sleep(0.5)
# ID統合 は いいえ
driver.find_element_by_xpath('//*[@id="area-confirm-dialog-button-cancel"]').click()
# いいえ の場合はアラートが表示されるのでクリックする
Alert(driver).accept()
print(my_function.browser_load_wait(driver))
# ------------------------------
# 受信トレイへ移動してメル確認
url = "https://m.kuku.lu/recv.php"
driver.get(url)
print(my_function.browser_load_wait(driver))
soup = BeautifulSoup(driver.page_source, "html.parser")

# for id_tag in soup.find_all(id=re.compile("area_mail_title_")):
# 	# "area_mail_title_" から始まって数字(桁数制限なし)で終わる範囲を切り取り
# 	id_tag = re.search('area_mail_title_\d+', str(id_tag)).group()
# 	print(id_tag)

time.sleep(1)
driver.find_element_by_xpath('//*[@id="area_mail_title_28710353"]').click()
time.sleep(3)
soup = BeautifulSoup(driver.page_source, "html.parser")
print(soup)

# for id_tag in soup.find_all(id=re.compile("area_maildata_")):
# 	# "area_maildata_" から始まって数字(桁数制限なし)で終わる範囲を切り取り
# 	id_tag = re.search('area_maildata_\d+', str(id_tag)).group()
# 	print(id_tag)
# id→area_maildata_28710353 を取得
# id→area-data を取得
# その中に件名とかコードが入っている

time.sleep(5)
driver.quit()