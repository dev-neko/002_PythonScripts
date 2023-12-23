'''
auidにクーポンが付与されているか確認するスクリプト

使い方
？

更新履歴
？

'''

import time
import re
from datetime import datetime
from selenium import webdriver
from my_module import my_function
import chromedriver_binary
from datetime import datetime
from my_module import mail_receiving_pop3
import poplib
import win32com
import xl
import openpyxl
import random
from my_module import trashmail_mass_production
from bs4 import BeautifulSoup
from ctypes import windll



# Excelの宣言・変数
Excel = 'aucoupon_data.xlsx'
wbpx = openpyxl.load_workbook(Excel)
wbxl = xl.Workbook(Excel)
# シート指定
sheet = wbpx['ID作成済み・付与確認']
# 付与確認開始行
row_str = 1463
# 付与確認終了行
row_end = 1573



# アカウントの数だけループさせる
print(row_str, "から", row_end, "まで実行")

for now_row in range(row_str, row_end+1):
# for now_row in range(5):
	driver = webdriver.Chrome()
	driver.set_window_size(1015, 515)

	# ログインページ開いてログイン
	url = "https://connect.auone.jp/net/vwc/cca_lg_eu_nets/login?targeturl=https%3A%2F%2Fid.auone.jp%2Findex.html%3Fstate%3Dlogin"
	driver.get(url)
	print(my_function.browser_load_wait(driver))
	auid_id   = sheet.cell(row=now_row, column=2).value
	auid_pass = sheet.cell(row=now_row, column=3).value
	print(auid_id)
	print(auid_pass)
	driver.find_element_by_xpath('//*[@id="loginAliasId"]').send_keys(auid_id)
	driver.find_element_by_xpath('//*[@id="loginAuonePwd"]').send_keys(auid_pass)
	driver.find_element_by_xpath('//*[@id="btn_login"]').click()
	time.sleep(2)

	# クーポンページを開いて付与確認
	url = "https://my.au.com/aus/WCV425001/WCE425001.hc?aa_bid=we-we-hd-9005"
	driver.get(url)
	print(my_function.browser_load_wait(driver))
	soup = BeautifulSoup(driver.page_source, "html.parser")
	if re.search('au乗りかえクーポン', str(soup)):
		print("クーポン 有り")
		wbxl.get("D"+str(now_row)).set("○")
	elif re.search('ご利用いただけるクーポンはありません', str(soup)):
		print("クーポン 無し")
		wbxl.get("D"+str(now_row)).set("× ")
	else:
		print("エラー")
		wbxl.get("D"+str(now_row)).set("エラー")

	driver.quit()

	print(now_row, "/", row_end, "付与確認完了")

# 正常終了したらスリープ状態に移行する
windll.powrprof.SetSuspendState(0, 0, 0)