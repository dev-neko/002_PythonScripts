'''
使い方

作りたいIDの数だけエクセルのセルを埋める
input_data タブをアクティブにする
トラメのメルアドを全て消す

更新履歴

名前とかの個人情報追加する部分は無い
browser_1に追加した

'''

import time
import re
from datetime import datetime
from selenium import webdriver
from my_module import my_function
import chromedriver_binary
from datetime import datetime
from my_module import mail_receiving_pop3
import poplib
import win32com
import xl
import openpyxl
import random
from my_module import trashmail_mass_production
from ctypes import windll



# POP3の宣言・変数
pop3_serv = 'pop.mail.yahoo.co.jp'
pop3_user = 'trashmail_receive_001'
pop3_pass = 'j50fBU1PvH'

# Excelの宣言・変数
Excel = 'aucoupon_data.xlsx'
wbpx = openpyxl.load_workbook(Excel)
wbxl = xl.Workbook(Excel)
# input_data
sheet = wbpx['input_data']
mrow = sheet.max_row
mcol = sheet.max_column

# トラメの転送先メルアド
trans_mail = 'trashmail_receive_001@yahoo.co.jp'



# メルアドの数だけループさせる
print(mrow, "まで実行")

for fa in range(mrow-1):
# for fa in range(3):
	# Tor経由で起動
	driver = my_function.chrome_connect_tor()
	driver.set_window_size(1015, 515)
	print("IPアドレス：" + my_function.Tor_IPaddr())

	# trashmail_mass_productionでメルアドを作成してExcelに記入
	mail_add = trashmail_mass_production.create_mail(driver, trans_mail)
	wbxl.get("A" + str(2 + fa)).set(mail_add)

	# POP3でエラーが出るが、再度実行すれば大丈夫か？
	while True:
		try:
			cli = poplib.POP3(pop3_serv)
			cli.user(pop3_user)
			cli.pass_(pop3_pass)
			# 確認コードが記載されているメルを受信する前のメルの数を取得
			mail_num_max = cli.stat()[0]
			print('受信前のメル数', mail_num_max)
		except:
			print('Error')
		else:
			break

	url = "https://connect.auone.jp/net/vw/cca_eu_net/cca?ID=ENET0510&targeturl=https%3a%2f%2fwww%2eau%2ecom%2fshopping%2fwowma%2flp%2fcashback%2dcmp%2f%3Faa_bid%3Did-cpn-001"
	driver.get(url)
	print(my_function.browser_load_wait(driver))

	# 作成したメルアドを入力
	driver.find_element_by_xpath('//*[@id="wowAliasIdEmail"]').send_keys(mail_add)
	driver.find_element_by_xpath('//*[@id="middleArea"]/div/form/div/div/div[1]/div/div/section/div[4]/button').click()
	print(my_function.browser_load_wait(driver))

	# 受信していないとエラーになるのでtryでループする
	mail_receive_count = 0
	while True:
		try:
			cli = poplib.POP3(pop3_serv)
			cli.user(pop3_user)
			cli.pass_(pop3_pass)
			msg = mail_receiving_pop3.fetchmail(cli, mail_num_max + 1)
		except poplib.error_proto as err:
			mail_receive_count += 1
			print(mail_receive_count, 'メル受信待ち', mail_num_max)
		else:
			if msg[2] == "【au ID】新規登録":
				print("受信したしauidからだった")
				break
			else:
				print("受信したけどauidからじゃなかった")
				mail_num_max += 1
		time.sleep(1)

	# 6文字の数字を取得
	conf_code = re.search("[0-9]{6}", msg[3])
	conf_code = conf_code.group()
	print(conf_code)

	# 確認コードを入力して「次へ」をクリック
	driver.find_element_by_xpath('//*[@id="confirmcode"]').send_keys(conf_code)
	driver.find_element_by_xpath('//*[@id="middleArea"]/div/form/div/div/div[1]/div/div/section/div[5]/button').click()
	print(my_function.browser_load_wait(driver))

	# パスワード、生年月日、性別を入力して新規登録を押す
	# パスワードは作成してExcelに記入
	# 8桁以上英数記号
	wowid_pass = my_function.pass_gen(20)
	# 作成したパスをエクセルに記入
	wbxl.get("B" + str(2 + fa)).set(wowid_pass)
	driver.find_element_by_xpath('//*[@id="password"]').send_keys(wowid_pass)
	# 生年月日性別は範囲指定のランダムで作成→1960～1999まで
	inp_year = "19" + str(random.randrange(60, 100))
	inp_mon = (random.randrange(1, 13))
	inp_day = (random.randrange(1, 29))
	driver.find_element_by_xpath('//*[@id="csBirthdayYYYY"]').send_keys(inp_year)
	driver.find_element_by_xpath('//*[@id="csBirthdayMM"]').send_keys(inp_mon)
	driver.find_element_by_xpath('//*[@id="csBirthdayDD"]').send_keys(inp_day)
	# 性別
	sex_num = random.randrange(2)
	if sex_num == 1:
		# 男
		sex_xpass = '//*[@id="middleArea"]/div/form/div/div/div[1]/div/div/section/div[7]/div/label[1]'
	else:
		# 女
		sex_xpass = '//*[@id="middleArea"]/div/form/div/div/div[1]/div/div/section/div[7]/div/label[2]'
	driver.find_element_by_xpath(sex_xpass).click()

	# 新規登録ボタンを押す
	driver.find_element_by_xpath('//*[@id="btn_cmp"]').click()
	print(my_function.browser_load_wait(driver))

	# 登録後の次へボタンを押す
	# なぜかここから先に進むのが遅いのでここは飛ばした
	# driver.find_element_by_xpath('//*[@id="middleArea"]/div/form/div/div/div[1]/div/div/section/div[6]/button').click()
	# time.sleep(2)

	print(fa+1, "/", mrow-1, "auid作成完了")

	driver.quit()

# 正常終了したらスリープ状態に移行する
windll.powrprof.SetSuspendState(0, 0, 0)